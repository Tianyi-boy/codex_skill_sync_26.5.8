from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROWS = [
    ("Amazon", "kagu", "Amazon Ads"),
    ("Amazon", "genji", "Amazon Ads"),
    ("Amazon", "senyu", "Amazon Ads"),
    ("Amazon", "zhongcheng", "Amazon Ads"),
    ("Wayfair", "linhe", "Wayfair Ads"),
    ("Shopify", "anzhap", "Google Ads"),
    ("Shopify", "anzhap", "Pinterest Ads"),
    ("Shopify", "anzhap", "Facebook Ads"),
    ("Shopify", "kagu", "Google Ads"),
    ("Shopify", "kagu", "Pinterest Ads"),
    ("Shopify", "kagu", "Facebook Ads"),
    ("Shopify", "zima", "Google Ads"),
    ("Shopify", "zima", "Pinterest Ads"),
    ("Shopify", "zima", "Facebook Ads"),
]
STORE_ALIASES = {
    "kaguyasu-us": "kagu",
    "kaguyasu": "kagu",
    "genji furniture-us": "genji",
    "genji furniture": "genji",
}
OBSERVE = "持续观察"
ADJUSTED = "已调整"
REJECTED = "拒绝调整"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a daily ad adjustment briefing.")
    parser.add_argument("--report-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--reports-dir", required=True, help="Directory containing generated reports.")
    parser.add_argument("--records-json", action="append", default=[], help="Validated/imported records JSON. Repeatable.")
    parser.add_argument("--records-dir", help="Optional directory containing records JSON files.")
    parser.add_argument("--performance-dir", help="Optional directory containing source ad performance workbooks.")
    parser.add_argument("--output-md", help="Optional Markdown output path.")
    parser.add_argument("--output-json", help="Optional JSON output path.")
    return parser.parse_args()


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                pass
    return None


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_platform(value: Any) -> str:
    text = cell_text(value).lower()
    if "amazon" in text or "亚马逊" in text:
        return "Amazon"
    if "wayfair" in text:
        return "Wayfair"
    if "shopify" in text:
        return "Shopify"
    return cell_text(value)


def normalize_store(value: Any) -> str:
    text = cell_text(value)
    key = text.lower()
    return STORE_ALIASES.get(key, key)


def normalize_channel(value: Any) -> str:
    text = cell_text(value)
    aliases = {
        "amazon ads": "Amazon Ads",
        "wayfair ads": "Wayfair Ads",
        "google ads": "Google Ads",
        "pinterest ads": "Pinterest Ads",
        "facebook ads": "Facebook Ads",
    }
    return aliases.get(text.lower(), text)


def canonical_key(platform: Any, store: Any, channel: Any) -> tuple[str, str, str] | None:
    key = (normalize_platform(platform), normalize_store(store), normalize_channel(channel))
    return key if key in ROWS else None


def header_map(values: list[Any]) -> dict[str, int]:
    aliases = {
        "platform": ("平台", "platform"),
        "store": ("店铺", "store"),
        "channel": ("投放渠道", "ad_channel", "channel"),
        "report_date": ("报告日期", "report_date"),
        "campaign_id": ("广告活动ID", "广告活动 ID", "campaign_id"),
        "campaign": ("广告活动名称", "campaign", "campaign_name"),
        "action": ("建议动作", "根据规则调整的动作", "recommended_action"),
        "status": ("是否调整", "adjustment_status"),
        "adjustment_time": ("调整时间", "adjustment_time"),
    }
    normalized = {cell_text(value).lower(): index for index, value in enumerate(values)}
    result: dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            if name.lower() in normalized:
                result[field] = normalized[name.lower()]
                break
    return result


def read_generated_report_key(path: Path, report_date: date) -> tuple[str, str, str] | None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]
    mapping = header_map(headers)
    required = {"platform", "store", "channel", "report_date"}
    if not required.issubset(mapping):
        return None
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        if parse_date(row[mapping["report_date"]]) == report_date:
            return canonical_key(row[mapping["platform"]], row[mapping["store"]], row[mapping["channel"]])
    return None


def generated_report_keys(reports_dir: Path, report_date: date) -> set[tuple[str, str, str]]:
    keys: set[tuple[str, str, str]] = set()
    for path in reports_dir.rglob("*.xlsx"):
        if path.name.startswith("~$") or report_date.isoformat() not in path.name:
            continue
        try:
            key = read_generated_report_key(path, report_date)
        except Exception:
            continue
        if key:
            keys.add(key)
    return keys


def records_paths(args: argparse.Namespace) -> list[Path]:
    paths = [Path(value) for value in args.records_json]
    if args.records_dir:
        paths.extend(Path(args.records_dir).rglob("*.json"))
    return [path for path in paths if path.exists()]


def unpack_records(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        records = data.get("records") or data.get("adjustment_records") or data.get("data")
        if isinstance(records, list):
            return [item for item in records if isinstance(item, dict)]
    return []


def record_field(record: dict[str, Any], key: str, raw_key: str | None = None) -> Any:
    if key in record:
        return record[key]
    raw = record.get("raw_report_row")
    if isinstance(raw, dict):
        return raw.get(raw_key or key)
    return None


def load_records(paths: list[Path], report_date: date) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for path in paths:
        data = json.loads(path.read_text(encoding="utf-8"))
        for record in unpack_records(data):
            if parse_date(record_field(record, "report_date", "报告日期")) != report_date:
                continue
            key = canonical_key(
                record_field(record, "platform", "平台"),
                record_field(record, "store", "店铺"),
                record_field(record, "ad_channel", "投放渠道") or record_field(record, "channel", "投放渠道"),
            )
            if not key:
                continue
            records.append(
                {
                    "key": key,
                    "recommended_action": cell_text(record_field(record, "recommended_action", "建议动作")),
                    "adjustment_status": cell_text(record_field(record, "adjustment_status", "是否调整")),
                    "adjustment_time": parse_date(record_field(record, "adjustment_time", "调整时间")),
                }
            )
    return records


def latest_performance_date(performance_dir: str | None) -> date | None:
    if not performance_dir:
        return None
    latest: date | None = None
    for path in Path(performance_dir).rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            headers = [cell_text(cell.value).lower() for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            date_indexes = [i for i, value in enumerate(headers) if value in {"date", "日期", "预估日期"}]
            if not date_indexes:
                continue
            date_index = date_indexes[0]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                parsed = parse_date(row[date_index])
                if parsed and (latest is None or parsed > latest):
                    latest = parsed
        except Exception:
            continue
    return latest


def aggregate_rows(got_reports: set[tuple[str, str, str]], records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        grouped[record["key"]].append(record)
    rows: list[dict[str, Any]] = []
    for platform, store, channel in ROWS:
        key = (platform, store, channel)
        row_records = grouped[key]
        need = [r for r in row_records if OBSERVE not in r["recommended_action"]]
        observe = [r for r in row_records if OBSERVE in r["recommended_action"]]
        rows.append(
            {
                "platform": platform,
                "store": store,
                "ad_channel": channel,
                "got_report": key in got_reports or bool(row_records),
                "submitted_report": bool(row_records),
                "submitted_rows": len(row_records),
                "need_adjustment": len(need),
                "continue_observation": len(observe),
                "completed_adjustment": sum(1 for r in need if r["adjustment_status"] == ADJUSTED),
                "rejected_adjustment": sum(1 for r in need if r["adjustment_status"] == REJECTED),
            }
        )
    return rows


def recent_effect_text(records: list[dict[str, Any]], report_date: date, latest_date: date | None) -> str:
    start = report_date - timedelta(days=13)
    recent = [r for r in records if r["adjustment_time"] and start <= r["adjustment_time"] <= report_date + timedelta(days=1)]
    if not recent:
        return "近14天没有可核验的已提交调整记录。"
    min_time = min(r["adjustment_time"] for r in recent if r["adjustment_time"])
    if latest_date is None:
        return f"近14天有 {len(recent)} 条调整记录，但未提供可解析的广告表现数据，不能评估调整效果。"
    if latest_date <= min_time:
        missing_start = min_time + timedelta(days=1)
        return (
            f"近14天有 {len(recent)} 条调整记录；调整时间最早为 {min_time.isoformat()}，"
            f"广告表现数据最新到 {latest_date.isoformat()}，缺少 {missing_start.isoformat()} 之后的后置观察数据，暂不能评估效果。"
        )
    return (
        f"近14天有 {len(recent)} 条调整记录，且广告表现数据已到 {latest_date.isoformat()}；"
        "可进入逐广告前后窗口对比，结论需以 Spend、Sales、Orders、ACoS、CTR、CVR、CPA、ROAS 的后置数据为准。"
    )


def render_markdown(report_date: date, rows: list[dict[str, Any]], effect: str) -> str:
    header = (
        "| 平台 | 店铺 | 投放渠道 | 今日获取调整报告 | 今日提交调整报告 | 提交报告条数 | "
        "需要调整 | 持续观察 | 已完成调整 | 拒绝调整 |"
    )
    sep = "|---|---|---|---:|---:|---:|---:|---:|---:|---:|"
    lines = [f"**广告调整简报｜{report_date.isoformat()}**", "", header, sep]
    for row in rows:
        lines.append(
            "| {platform} | {store} | {ad_channel} | {got} | {submitted} | {submitted_rows} | "
            "{need_adjustment} | {continue_observation} | {completed_adjustment} | {rejected_adjustment} |".format(
                **row,
                got="是" if row["got_report"] else "否",
                submitted="是" if row["submitted_report"] else "否",
            )
        )
    total = {
        "stores": sum(1 for row in rows if row["submitted_report"]),
        "submitted": sum(row["submitted_rows"] for row in rows),
        "need": sum(row["need_adjustment"] for row in rows),
        "observe": sum(row["continue_observation"] for row in rows),
        "done": sum(row["completed_adjustment"] for row in rows),
        "reject": sum(row["rejected_adjustment"] for row in rows),
    }
    lines.extend(
        [
            "",
            f"今日共确认 {total['stores']} 个店铺/渠道提交调整报告，提交 {total['submitted']} 条；"
            f"其中 {total['need']} 条需要调整、{total['observe']} 条持续观察，"
            f"需要调整中 {total['done']} 条已完成、{total['reject']} 条拒绝调整。",
            "",
            "**近14天调整效果**",
            "",
            effect,
        ]
    )
    return "\n".join(lines)


def main() -> int:
    args = parse_args()
    report_date = datetime.strptime(args.report_date, "%Y-%m-%d").date()
    got_reports = generated_report_keys(Path(args.reports_dir), report_date)
    records = load_records(records_paths(args), report_date)
    rows = aggregate_rows(got_reports, records)
    effect = recent_effect_text(records, report_date, latest_performance_date(args.performance_dir))
    payload = {"report_date": report_date.isoformat(), "rows": rows, "recent_effect": effect}
    markdown = render_markdown(report_date, rows, effect)
    if args.output_json:
        Path(args.output_json).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.output_md:
        Path(args.output_md).write_text(markdown, encoding="utf-8")
    print(markdown)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
