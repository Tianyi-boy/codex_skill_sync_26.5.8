from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "campaign_identity": ("广告活动ID", "广告活动 ID", "广告活动名称", "广告活动", "campaign_id", "Campaign ID"),
    "campaign_name": ("广告活动名称", "广告活动", "campaign_name", "Campaign Name"),
    "recommended_action": ("建议动作", "寤议动作", "recommended_action"),
    "triggered_rule": ("触发规则", "triggered_rule"),
    "adjust_status": ("是否调整",),
    "adjust_method": ("调整方式",),
    "reject_reason": ("拒绝调整原因",),
    "adjusted_at": ("调整时间",),
}

METRIC_PREFIXES = (
    "ACoS_",
    "CVR_",
    "CTR_",
    "Spend_",
    "Sales_",
    "Orders_",
    "Clicks_",
    "Impressions_",
    "spend_",
    "sales_",
    "orders_",
    "clicks_",
    "impressions_",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate a completed ad adjustment report.")
    parser.add_argument("--report", required=True, help="Path to completed report workbook.")
    return parser.parse_args()


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_date(value: Any) -> bool:
    if value is None:
        return False
    if hasattr(value, "date"):
        return True
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            datetime.strptime(text[:19], fmt)
            return True
        except ValueError:
            try:
                datetime.strptime(text[:10], fmt)
                return True
            except ValueError:
                pass
    return False


def raw_header_map(worksheet: Any) -> dict[str, int]:
    return {
        str(worksheet.cell(2, col).value).strip(): col
        for col in range(1, worksheet.max_column + 1)
        if worksheet.cell(2, col).value
    }


def find_header(mapping: dict[str, int], aliases: tuple[str, ...]) -> int | None:
    for alias in aliases:
        if alias in mapping:
            return mapping[alias]
    return None


def required_header_map(worksheet: Any) -> dict[str, int | list[int]]:
    raw = raw_header_map(worksheet)
    resolved: dict[str, int | list[int]] = {}
    missing: list[str] = []
    for key, aliases in HEADER_ALIASES.items():
        if key == "campaign_identity":
            cols = [raw[alias] for alias in aliases if alias in raw]
            if not cols:
                missing.append(f"{key}: one of {list(aliases)}")
            else:
                resolved[key] = cols
            continue
        col = find_header(raw, aliases)
        if col is None:
            missing.append(f"{key}: one of {list(aliases)}")
        else:
            resolved[key] = col
    if missing:
        raise ValueError(f"REPORT_SCHEMA_MISMATCH: missing headers {missing}")
    return resolved


def metric_columns(worksheet: Any) -> list[int]:
    raw = raw_header_map(worksheet)
    return [col for header, col in raw.items() if header.startswith(METRIC_PREFIXES) or "空白原因" in header]


def has_identity(worksheet: Any, row_index: int, mapping: dict[str, int | list[int]]) -> bool:
    identity_cols = mapping["campaign_identity"]
    assert isinstance(identity_cols, list)
    return any(not is_blank(worksheet.cell(row_index, col).value) for col in identity_cols)


def cell_value(worksheet: Any, row_index: int, mapping: dict[str, int | list[int]], key: str) -> Any:
    col = mapping[key]
    assert isinstance(col, int)
    return worksheet.cell(row_index, col).value


def validate(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    mapping = required_header_map(worksheet)
    metrics = metric_columns(worksheet)
    failures: list[dict[str, Any]] = []
    checked_rows = 0

    for row_index in range(3, worksheet.max_row + 1):
        if not has_identity(worksheet, row_index, mapping):
            continue
        checked_rows += 1
        status = cell_value(worksheet, row_index, mapping, "adjust_status")
        method = cell_value(worksheet, row_index, mapping, "adjust_method")
        reject_reason = cell_value(worksheet, row_index, mapping, "reject_reason")
        adjusted_at = cell_value(worksheet, row_index, mapping, "adjusted_at")

        if status not in {"已调整", "拒绝调整"}:
            failures.append({"row": row_index, "field": "是否调整", "message": "must be 已调整 or 拒绝调整"})
        if status == "已调整" and is_blank(method):
            failures.append({"row": row_index, "field": "调整方式", "message": "required when 是否调整 is 已调整"})
        if status == "拒绝调整" and is_blank(reject_reason):
            failures.append({"row": row_index, "field": "拒绝调整原因", "message": "required when 是否调整 is 拒绝调整"})
        if not parse_date(adjusted_at):
            failures.append({"row": row_index, "field": "调整时间", "message": "required and must be parseable as a date"})

        for key, field_name in [
            ("recommended_action", "建议动作"),
            ("triggered_rule", "触发规则"),
        ]:
            if is_blank(cell_value(worksheet, row_index, mapping, key)):
                failures.append({"row": row_index, "field": field_name, "message": "required identity/action/rule field is blank"})

        if not any(not is_blank(worksheet.cell(row_index, col).value) for col in metrics):
            failures.append({"row": row_index, "field": "metric evidence", "message": "at least one metric evidence field must be present"})

    return {
        "status": "success" if not failures else "error",
        "checked_rows": checked_rows,
        "failure_count": len(failures),
        "failures": failures,
    }


def main() -> int:
    result = validate(Path(parse_args().report))
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0 if result["status"] == "success" else 2


if __name__ == "__main__":
    raise SystemExit(main())
