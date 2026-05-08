from __future__ import annotations

import argparse
import hashlib
import importlib.util
import json
import subprocess
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_CONTAINER = "agent-platform-mysql"
DEFAULT_DATABASE = "agent_platform"
DEFAULT_USER = "agent_app"
DEFAULT_PASSWORD = "change_me"
SKILL_ID = "ad-adjustment-report-lifecycle"


BASIC_FIELDS = [
    "平台",
    "店铺",
    "投放渠道",
    "报告日期",
    "数据周期",
    "店铺名称",
    "国家",
    "类型",
    "广告组合",
    "广告活动ID",
    "isB2b",
    "定位类型",
    "广告活动名称",
    "每日预算",
    "总预算",
    "开始日期",
    "结束日期",
    "竞价策略",
    "目标ROAS",
    "店铺网址",
]
REQUIRED_FIELDS = ["店铺", "报告日期", "类型", "广告活动名称", "建议动作", "触发规则", "是否调整", "调整方式", "拒绝调整原因", "调整时间"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import a validated ad adjustment report into Docker MySQL.")
    parser.add_argument("--report", required=True, help="Path to a completed adjustment report workbook.")
    parser.add_argument("--container", default=DEFAULT_CONTAINER, help=f"Docker MySQL container name. Default: {DEFAULT_CONTAINER}.")
    parser.add_argument("--database", default=DEFAULT_DATABASE, help=f"MySQL database name. Default: {DEFAULT_DATABASE}.")
    parser.add_argument("--user", default=DEFAULT_USER, help=f"MySQL user. Default: {DEFAULT_USER}.")
    parser.add_argument("--password", default=DEFAULT_PASSWORD, help="MySQL password.")
    parser.add_argument("--request-id", default="manual-local")
    parser.add_argument("--conversation-id", default="codex-desktop")
    parser.add_argument("--task-id", default="ad-adjustment-report-import")
    parser.add_argument("--agent-id", default="codex")
    parser.add_argument("--dry-run", action="store_true", help="Validate and summarize without writing to MySQL.")
    return parser.parse_args()


def load_validator() -> Any:
    path = Path(__file__).with_name("validate_completed_report.py")
    spec = importlib.util.spec_from_file_location("validate_completed_report", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load validator from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def validate_report(path: Path) -> dict[str, Any]:
    result = load_validator().validate(path)
    if result["status"] != "success":
        print(json.dumps(result, ensure_ascii=False, indent=2))
        raise SystemExit(2)
    return result


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def as_date(value: Any, field: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date().isoformat()
        except ValueError:
            try:
                return datetime.strptime(text[:10], fmt).date().isoformat()
            except ValueError:
                pass
    raise ValueError(f"{field} is not parseable as a date: {value!r}")


def cell_json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return value


def header_map(worksheet: Any) -> dict[str, int]:
    return {
        str(worksheet.cell(2, col).value).strip(): col
        for col in range(1, worksheet.max_column + 1)
        if worksheet.cell(2, col).value
    }


def require_headers(mapping: dict[str, int]) -> None:
    missing = [field for field in REQUIRED_FIELDS if field not in mapping]
    if missing:
        raise ValueError(f"REPORT_SCHEMA_MISMATCH: missing headers {missing}")


def stable_hash(*parts: Any) -> str:
    text = "\u241f".join("" if part is None else str(part) for part in parts)
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def read_rows(report: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(report, data_only=True)
    worksheet = workbook.active
    mapping = header_map(worksheet)
    require_headers(mapping)
    rows: list[dict[str, Any]] = []
    for row_index in range(3, worksheet.max_row + 1):
        campaign = worksheet.cell(row_index, mapping["广告活动名称"]).value
        campaign_id = worksheet.cell(row_index, mapping["广告活动ID"]).value if "广告活动ID" in mapping else ""
        if is_blank(campaign) and is_blank(campaign_id):
            continue
        raw = {
            str(worksheet.cell(2, col).value): cell_json_value(worksheet.cell(row_index, col).value)
            for col in range(1, worksheet.max_column + 1)
            if worksheet.cell(2, col).value
        }
        report_date = as_date(worksheet.cell(row_index, mapping["报告日期"]).value, "报告日期")
        store = str(worksheet.cell(row_index, mapping["店铺"]).value).strip()
        ad_type = str(worksheet.cell(row_index, mapping["类型"]).value or "").strip()
        portfolio = worksheet.cell(row_index, mapping["广告组合"]).value if "广告组合" in mapping else None
        campaign_value = str(campaign or campaign_id).strip()
        recommended_action = str(worksheet.cell(row_index, mapping["建议动作"]).value or "").strip()
        triggered_rule = str(worksheet.cell(row_index, mapping["触发规则"]).value or "").strip()
        adjustment_status = str(worksheet.cell(row_index, mapping["是否调整"]).value or "").strip()
        adjustment_method = str(worksheet.cell(row_index, mapping["调整方式"]).value or "").strip()
        reject_reason = worksheet.cell(row_index, mapping["拒绝调整原因"]).value
        adjustment_time = as_date(worksheet.cell(row_index, mapping["调整时间"]).value, "调整时间")
        record_id = stable_hash(report_date, store, ad_type, portfolio, campaign_value, row_index, recommended_action, triggered_rule)
        rows.append(
            {
                "record_id": record_id,
                "report_date": report_date,
                "source_excel_row": row_index,
                "store": store,
                "ad_type": ad_type or "unknown",
                "portfolio": portfolio,
                "campaign": campaign_value,
                "effective_status": "active_in_source",
                "recommended_action": recommended_action,
                "triggered_rule": triggered_rule,
                "adjustment_status": adjustment_status,
                "adjustment_method": adjustment_method,
                "reject_reason": reject_reason,
                "adjustment_time": adjustment_time,
                "raw_report_row": raw,
            }
        )
    return rows


def sql_string(value: Any) -> str:
    if value is None:
        return "NULL"
    text = str(value)
    return "'" + text.replace("\\", "\\\\").replace("'", "''") + "'"


def sql_int(value: int) -> str:
    return str(int(value))


def sql_json(value: Any) -> str:
    return sql_string(json.dumps(value, ensure_ascii=False, separators=(",", ":")))


def create_schema_sql() -> str:
    return """
CREATE TABLE IF NOT EXISTS `ad_adjustment_import_batches` (
  `batch_id` varchar(64) NOT NULL,
  `request_id` varchar(64) NOT NULL,
  `conversation_id` varchar(64) NOT NULL,
  `task_id` varchar(128) NOT NULL,
  `agent_id` varchar(128) NOT NULL,
  `skill_id` varchar(128) NOT NULL,
  `source_report_file` varchar(1024) NOT NULL,
  `report_date` date NOT NULL,
  `store` varchar(128) NOT NULL,
  `status` varchar(32) NOT NULL,
  `record_count` int NOT NULL DEFAULT '0',
  `inserted_count` int NOT NULL DEFAULT '0',
  `skipped_count` int NOT NULL DEFAULT '0',
  `error_code` varchar(128) DEFAULT NULL,
  `error_message` text,
  `created_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
  PRIMARY KEY (`batch_id`),
  KEY `idx_ad_adjustment_import_batches_request` (`request_id`),
  KEY `idx_ad_adjustment_import_batches_report` (`report_date`,`store`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci;

CREATE TABLE IF NOT EXISTS `ad_adjustment_records` (
  `id` bigint NOT NULL AUTO_INCREMENT,
  `record_id` varchar(64) NOT NULL,
  `batch_id` varchar(64) NOT NULL,
  `request_id` varchar(64) NOT NULL,
  `conversation_id` varchar(64) NOT NULL,
  `task_id` varchar(128) NOT NULL,
  `agent_id` varchar(128) NOT NULL,
  `skill_id` varchar(128) NOT NULL,
  `report_date` date NOT NULL,
  `source_excel_row` int NOT NULL,
  `store` varchar(128) NOT NULL,
  `ad_type` varchar(64) NOT NULL,
  `portfolio` varchar(255) DEFAULT NULL,
  `campaign` varchar(255) NOT NULL,
  `effective_status` varchar(64) NOT NULL,
  `recommended_action` text NOT NULL,
  `triggered_rule` text NOT NULL,
  `adjustment_status` varchar(32) NOT NULL,
  `adjustment_method` text NOT NULL,
  `reject_reason` text,
  `adjustment_time` date NOT NULL,
  `source_report_file` varchar(1024) NOT NULL,
  `raw_report_row` json NOT NULL,
  `created_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP,
  `updated_at` timestamp NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  PRIMARY KEY (`id`),
  UNIQUE KEY `uq_ad_adjustment_records_record_id` (`record_id`),
  KEY `idx_ad_adjustment_records_batch` (`batch_id`),
  KEY `idx_ad_adjustment_records_report` (`report_date`,`store`),
  KEY `idx_ad_adjustment_records_campaign` (`store`,`campaign`),
  CONSTRAINT `fk_ad_adjustment_records_batch` FOREIGN KEY (`batch_id`) REFERENCES `ad_adjustment_import_batches` (`batch_id`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci;
""".strip()


def build_import_sql(args: argparse.Namespace, report: Path, rows: list[dict[str, Any]]) -> tuple[str, str]:
    now = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S%f")
    report_date = rows[0]["report_date"] if rows else "1970-01-01"
    store = rows[0]["store"] if rows else ""
    batch_id = stable_hash("batch", report, now)
    values = []
    for row in rows:
        values.append(
            "("
            + ",".join(
                [
                    sql_string(row["record_id"]),
                    sql_string(report_date),
                    sql_int(row["source_excel_row"]),
                    sql_string(row["store"]),
                    sql_string(row["ad_type"]),
                    sql_string(row["portfolio"]),
                    sql_string(row["campaign"]),
                    sql_string(row["effective_status"]),
                    sql_string(row["recommended_action"]),
                    sql_string(row["triggered_rule"]),
                    sql_string(row["adjustment_status"]),
                    sql_string(row["adjustment_method"]),
                    sql_string(row["reject_reason"]),
                    sql_string(row["adjustment_time"]),
                    sql_json(row["raw_report_row"]),
                ]
            )
            + ")"
        )
    stage_insert = ""
    if values:
        stage_insert = """
INSERT INTO `ad_adjustment_records_stage`
  (`record_id`,`report_date`,`source_excel_row`,`store`,`ad_type`,`portfolio`,`campaign`,`effective_status`,`recommended_action`,`triggered_rule`,`adjustment_status`,`adjustment_method`,`reject_reason`,`adjustment_time`,`raw_report_row`)
VALUES
""".strip() + "\n" + ",\n".join(values) + ";"

    sql = f"""
SET NAMES utf8mb4;
START TRANSACTION;
{create_schema_sql()}
DROP TEMPORARY TABLE IF EXISTS `ad_adjustment_records_stage`;
CREATE TEMPORARY TABLE `ad_adjustment_records_stage` (
  `record_id` varchar(64) NOT NULL,
  `report_date` date NOT NULL,
  `source_excel_row` int NOT NULL,
  `store` varchar(128) NOT NULL,
  `ad_type` varchar(64) NOT NULL,
  `portfolio` varchar(255) DEFAULT NULL,
  `campaign` varchar(255) NOT NULL,
  `effective_status` varchar(64) NOT NULL,
  `recommended_action` text NOT NULL,
  `triggered_rule` text NOT NULL,
  `adjustment_status` varchar(32) NOT NULL,
  `adjustment_method` text NOT NULL,
  `reject_reason` text,
  `adjustment_time` date NOT NULL,
  `raw_report_row` json NOT NULL,
  PRIMARY KEY (`record_id`)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci;
{stage_insert}
SET @record_count := (SELECT COUNT(*) FROM `ad_adjustment_records_stage`);
INSERT INTO `ad_adjustment_import_batches`
  (`batch_id`,`request_id`,`conversation_id`,`task_id`,`agent_id`,`skill_id`,`source_report_file`,`report_date`,`store`,`status`,`record_count`)
VALUES
  ({sql_string(batch_id)},{sql_string(args.request_id)},{sql_string(args.conversation_id)},{sql_string(args.task_id)},{sql_string(args.agent_id)},{sql_string(SKILL_ID)},{sql_string(str(report))},{sql_string(report_date)},{sql_string(store)},'running',@record_count);
INSERT IGNORE INTO `ad_adjustment_records`
  (`record_id`,`batch_id`,`request_id`,`conversation_id`,`task_id`,`agent_id`,`skill_id`,`report_date`,`source_excel_row`,`store`,`ad_type`,`portfolio`,`campaign`,`effective_status`,`recommended_action`,`triggered_rule`,`adjustment_status`,`adjustment_method`,`reject_reason`,`adjustment_time`,`source_report_file`,`raw_report_row`)
SELECT
  `record_id`,{sql_string(batch_id)},{sql_string(args.request_id)},{sql_string(args.conversation_id)},{sql_string(args.task_id)},{sql_string(args.agent_id)},{sql_string(SKILL_ID)},`report_date`,`source_excel_row`,`store`,`ad_type`,`portfolio`,`campaign`,`effective_status`,`recommended_action`,`triggered_rule`,`adjustment_status`,`adjustment_method`,`reject_reason`,`adjustment_time`,{sql_string(str(report))},`raw_report_row`
FROM `ad_adjustment_records_stage`;
SET @inserted_count := ROW_COUNT();
UPDATE `ad_adjustment_import_batches`
SET
  `status` = 'success',
  `inserted_count` = @inserted_count,
  `skipped_count` = @record_count - @inserted_count
WHERE `batch_id` = {sql_string(batch_id)};
COMMIT;
SELECT JSON_OBJECT(
  'status','success',
  'batch_id',`batch_id`,
  'record_count',`record_count`,
  'inserted_count',`inserted_count`,
  'skipped_count',`skipped_count`,
  'source_report_file',`source_report_file`,
  'report_date',DATE_FORMAT(`report_date`,'%Y-%m-%d'),
  'store',`store`
) FROM `ad_adjustment_import_batches` WHERE `batch_id` = {sql_string(batch_id)};
""".strip()
    return batch_id, sql


def run_mysql(args: argparse.Namespace, sql: str) -> dict[str, Any]:
    command = [
        "docker",
        "exec",
        "-i",
        args.container,
        "mysql",
        f"-u{args.user}",
        f"-p{args.password}",
        "--default-character-set=utf8mb4",
        "--batch",
        "--raw",
        "--skip-column-names",
        args.database,
    ]
    completed = subprocess.run(command, input=sql, text=True, capture_output=True)
    if completed.returncode != 0:
        print(completed.stderr.strip(), file=sys.stderr)
        raise SystemExit(completed.returncode)
    lines = [line for line in completed.stdout.splitlines() if line.strip()]
    if not lines:
        raise RuntimeError("Import SQL completed without returning a summary")
    return json.loads(lines[-1])


def check_mysql_connection(args: argparse.Namespace) -> dict[str, Any]:
    return run_mysql(
        args,
        "SELECT JSON_OBJECT('status','success','database',DATABASE(),'user',CURRENT_USER());",
    )


def main() -> int:
    args = parse_args()
    report = Path(args.report).resolve()
    validation = validate_report(report)
    rows = read_rows(report)
    if args.dry_run:
        connection = check_mysql_connection(args)
        print(
            json.dumps(
                {
                    "status": "success",
                    "dry_run": True,
                    "validation": validation,
                    "record_count": len(rows),
                    "source_report_file": str(report),
                    "mysql_connection": connection,
                    "mysql_target": {
                        "container": args.container,
                        "database": args.database,
                        "user": args.user,
                    },
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0
    _, sql = build_import_sql(args, report, rows)
    summary = run_mysql(args, sql)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
